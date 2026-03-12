import csv
import docx2txt
import faiss
import io
import logging
import os
import openpyxl
import PyPDF2
import re
from sentence_transformers import SentenceTransformer

class DocumentStore:
    """Handles document loading, chunking, and retrieval"""

    def __init__(self, embedding_model="all-MiniLM-L6-v2"):
        """Initialize document store with embeddings"""
        logging.info("Loading embedding model...")
        self.embedder = SentenceTransformer(embedding_model)
        self.chunks = []
        self.metadata = []
        self.index = None

    def load_document(self, file_path):
        """Load document from various formats"""
        ext = os.path.splitext(file_path)[1].lower()

        if ext == '.pdf':
            text = self._load_pdf(file_path)
        elif ext == '.docx':
            text = self._load_docx(file_path)
        elif ext == '.xlsx' or ext == '.xls':
            text = self._load_xlsx(file_path)
        elif ext == '.csv':
            text = self._load_csv(file_path)
        elif ext == '.txt':
            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()
        else:
            raise ValueError(f"Unsupported file format: {ext}")

        # Split into chunks
        chunks = self._chunk_text(text)

        for chunk in chunks:
            self.chunks.append(chunk)
            self.metadata.append({'source': os.path.basename(file_path)})

        logging.info(f"Loaded {len(chunks)} chunks from {os.path.basename(file_path)}")

    def _load_pdf(self, file_path):
        """Extract text from PDF"""
        text = ""
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                text += page.extract_text() + "\n"
        return text

    def _load_docx(self, file_path):
        """Extract text from DOCX"""
        with open(file_path, 'r'):
            text = docx2txt.process(file_path)
        return text

    def _load_csv(self, file_path):
        """Extract text from CSV"""
        with open(file_path, 'rb') as file:
            file.seek(0)
            content = file.read().decode('utf-8')
            csv_reader = csv.reader(io.StringIO(content))
            rows = list(csv_reader)

        return '\n'.join([', '.join(row) for row in rows])

    def _load_xlsx(self, file_path):
        """Extract text from Excel files (XLSX, XLS)"""
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        text_parts = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")

            # Get all rows
            for row in sheet.iter_rows(values_only=True):
                # Filter out None values and convert to strings
                row_values = [str(cell) for cell in row if cell is not None]
                if row_values:  # Only add non-empty rows
                    logging.error(row_values)
                    text_parts.append(" | ".join(row_values)+"\n")

        return "\n".join(text_parts)

    def _chunk_text(self, text, chunk_size=1000, overlap=50):
        """Split text into overlapping chunks"""
        # Clean text
        text = re.sub(r'\s+', ' ', text).strip()

        words = text.split()
        chunks = []

        for i in range(0, len(words), chunk_size - overlap):
            chunk = ' '.join(words[i:i + chunk_size])
            if chunk:
                chunks.append(chunk+"\n")

        return chunks

    def build_index(self):
        """Build FAISS index for similarity search"""
        if not self.chunks:
            logging.info("No documents loaded!")
            return

        logging.info("Building search index...")
        embeddings = self.embedder.encode(self.chunks, show_progress_bar=True)

        # Create FAISS index
        dimension = embeddings.shape[1]
        self.index = faiss.IndexFlatL2(dimension)
        self.index.add(embeddings.astype('float32'))
        logging.info(f"Index built with {len(self.chunks)} chunks\n")

    def search(self, query, k=3):
        """Search for relevant chunks"""
        if self.index is None:
            return []

        query_embedding = self.embedder.encode([query])
        distances, indices = self.index.search(query_embedding.astype('float32'), k)

        results = []
        for idx, dist in zip(indices[0], distances[0]):
            if idx < len(self.chunks):
                results.append({
                    'text': self.chunks[idx],
                    'source': self.metadata[idx]['source'],
                    'score': float(dist)
                })

        return results
